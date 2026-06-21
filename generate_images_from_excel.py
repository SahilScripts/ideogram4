#!/usr/bin/env python3
"""Generate images from an Excel sheet using the same Ideogram 4 flow as run_inference.py."""

from __future__ import annotations

import logging
import os
import sys
from pathlib import Path

import torch
from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent
SRC_DIR = BASE_DIR / "src"
if SRC_DIR.exists():
    sys.path.insert(0, str(SRC_DIR))

from ideogram4 import (  # noqa: E402
    DEFAULT_MAGIC_PROMPT,
    MAGIC_PROMPTS,
    PRESETS,
    Ideogram4Pipeline,
    Ideogram4PipelineConfig,
    aspect_ratio_from_size,
)

# ----------------- Configuration -----------------
# Hardcoded Ideogram API key used for hosted magic-prompt expansion.
IDEOGRAM_API_KEY = "eLJsiusE8uH6QkOFWzNe-aEcJaPtY-4f3E2YCE1GvrQYtZDmP71LXaMj3f4BpL4uyt00V9geJFNRZCrKQ56AYw"
# Excel file name in the same folder as this script.
EXCEL_FILENAME = "scenes.xlsx"
# Skip image generation if the output file already exists.
SKIP_EXISTING = True
# Inference settings mirrored from run_inference.py.
OUTPUT_WIDTH = 1024
OUTPUT_HEIGHT = 1024
SAMPLER_PRESET = "V4_QUALITY_48"
SEED = 0
USE_MAGIC_PROMPT = True
MAGIC_PROMPT_MODEL = DEFAULT_MAGIC_PROMPT
WARN_ON_CAPTION_ISSUES = False
# -------------------------------------------------

QUANTIZATION_REPOS = {
    "nf4": "ideogram-ai/ideogram-4-nf4",
    "fp8": "ideogram-ai/ideogram-4-fp8",
}

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s: %(message)s")


def _default_device() -> str:
    if torch.cuda.is_available():
        return "cuda"
    if torch.backends.mps.is_available():
        return "mps"
    return "cpu"


def _default_quantization() -> str:
    return "nf4" if torch.cuda.is_available() else "fp8"


def safe_filename(name: str) -> str:
    if not isinstance(name, str):
        name = str(name)
    for ch in ('/', "\\", ':', '*', '?', '"', '<', '>', '|'):
        name = name.replace(ch, '_')
    name = name.strip()
    return name or "unnamed_scene"


def _looks_like_header(row: tuple) -> bool:
    labels = []
    for index in range(min(3, len(row))):
        value = row[index]
        labels.append(str(value).strip().lower() if value is not None else "")
    header_tokens = {"scene", "scene name", "scene number", "timestamp", "timestamp range", "narration"}
    return any(label in header_tokens for label in labels)


def _resolve_excel_path() -> Path:
    return BASE_DIR / EXCEL_FILENAME


def _load_pipeline() -> Ideogram4Pipeline:
    return Ideogram4Pipeline.from_pretrained(
        config=Ideogram4PipelineConfig(weights_repo=QUANTIZATION_REPOS[_default_quantization()]),
        device=_default_device(),
        dtype=torch.bfloat16,
    )


def _expand_prompt(prompt: str, width: int, height: int) -> str:
    if not USE_MAGIC_PROMPT:
        return prompt
    if not IDEOGRAM_API_KEY or IDEOGRAM_API_KEY.startswith("REPLACE_"):
        raise RuntimeError("IDEOGRAM_API_KEY is not configured for magic-prompt expansion")
    aspect_ratio = aspect_ratio_from_size(width, height)
    magic_prompt = MAGIC_PROMPTS[MAGIC_PROMPT_MODEL](api_key=IDEOGRAM_API_KEY)  # type: ignore[call-arg]
    logging.info("Expanding prompt with %s for aspect ratio %s", MAGIC_PROMPT_MODEL, aspect_ratio)
    return magic_prompt.expand(prompt, aspect_ratio=aspect_ratio)


def _generate_images(pipe: Ideogram4Pipeline, prompt: str):
    preset = PRESETS[SAMPLER_PRESET]
    return pipe(
        prompt,
        height=OUTPUT_HEIGHT,
        width=OUTPUT_WIDTH,
        num_steps=preset.num_steps,
        guidance_schedule=preset.guidance_schedule,
        mu=preset.mu,
        std=preset.std,
        seed=SEED,
        raise_on_caption_issues=not WARN_ON_CAPTION_ISSUES,
    )


def main() -> None:
    if USE_MAGIC_PROMPT and (not IDEOGRAM_API_KEY or IDEOGRAM_API_KEY.startswith("REPLACE_")):
        logging.error("Please set IDEOGRAM_API_KEY at the top of %s", __file__)
        sys.exit(1)

    excel_path = _resolve_excel_path()
    if not excel_path.exists():
        logging.error("Excel file not found: %s", excel_path)
        sys.exit(1)

    workbook = load_workbook(excel_path, read_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        logging.error("Excel file is empty.")
        sys.exit(1)

    start_row_index = 1 if _looks_like_header(rows[0]) else 0
    data_rows = rows[start_row_index:]

    max_prompt_col = 3
    for row in data_rows:
        for col_index in range(3, len(row)):
            value = row[col_index]
            if value is not None and str(value).strip() != "":
                max_prompt_col = max(max_prompt_col, col_index + 1)

    if max_prompt_col <= 3:
        logging.error("No prompt columns detected (need at least column D).")
        sys.exit(1)

    num_prompt_cols = max_prompt_col - 3

    prompt_folders: list[Path] = []
    for index in range(num_prompt_cols):
        folder = BASE_DIR / f"prompt {index + 1}"
        folder.mkdir(parents=True, exist_ok=True)
        prompt_folders.append(folder)

    logging.info("Detected %d prompt columns. Output folders: %s", num_prompt_cols, ", ".join(str(folder) for folder in prompt_folders))

    pipe = _load_pipeline()

    for row_number, row in enumerate(data_rows, start=start_row_index + 1):
        scene_raw = row[0] if len(row) > 0 else None
        if scene_raw is None or str(scene_raw).strip() == "":
            continue

        scene_name = safe_filename(scene_raw)

        for prompt_index in range(num_prompt_cols):
            prompt_value = row[3 + prompt_index] if len(row) > 3 + prompt_index else None
            if prompt_value is None or str(prompt_value).strip() == "":
                logging.info('Row %d: "%s": Prompt column %d is empty - skipping', row_number, scene_name, prompt_index + 1)
                continue

            folder = prompt_folders[prompt_index]
            out_path = folder / f"{scene_name}.png"

            if SKIP_EXISTING and out_path.exists():
                logging.info('Row %d: "%s": File already exists in %s - skipping', row_number, scene_name, folder)
                continue

            logging.info('Row %d: "%s": Generating image for prompt %d...', row_number, scene_name, prompt_index + 1)
            try:
                expanded_prompt = _expand_prompt(str(prompt_value), OUTPUT_WIDTH, OUTPUT_HEIGHT)
                images = _generate_images(pipe, expanded_prompt)
                if images:
                    images[0].save(out_path)
                    logging.info('Row %d: "%s": Saved image to %s', row_number, scene_name, out_path)
                else:
                    logging.error('Row %d: "%s": Image generation returned no images for prompt %d', row_number, scene_name, prompt_index + 1)
            except Exception as exc:
                logging.error('Row %d: "%s": Image generation failed for prompt %d: %s', row_number, scene_name, prompt_index + 1, exc)

    logging.info("Processing complete.")


if __name__ == "__main__":
    main()
